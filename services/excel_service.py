"""
services/excel_service.py
Genera reportes Excel institucionales con openpyxl.
Estilo: cabecera verde Municipalidad de Yau, colores por prioridad/estado.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colores institucionales ────────────────────────────────────────
VERDE       = 'FF2E7D4F'
VERDE_OSC   = 'FF1B5E35'
VERDE_CLR   = 'FFE1F5EE'
BLANCO      = 'FFFFFFFF'
GRIS_HD     = 'FF444441'
GRIS_ALT    = 'FFF1EFE8'

COLORES_PRIORIDAD = {
    'critico': 'FFFCEBEB',
    'normal':  'FFFAEEDA',
    'bajo':    'FFE1F5EE',
}
COLORES_ESTADO = {
    'aprobado':  'FFE1F5EE',
    'rechazado': 'FFFCEBEB',
    'en_proceso':'FFE6F1FB',
    'pendiente': 'FFFAEEDA',
}


def _borde(grosor='thin'):
    s = Side(style=grosor)
    return Border(left=s, right=s, top=s, bottom=s)


def _celda(ws, fila, col, valor, negrita=False, color_font=None,
           color_fill=None, alineacion='left', tamanio=10, wrap=False):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font      = Font(name='Calibri', bold=negrita,
                       size=tamanio,
                       color=color_font or GRIS_HD)
    if color_fill:
        c.fill  = PatternFill('solid', fgColor=color_fill)
    c.alignment = Alignment(horizontal=alineacion,
                            vertical='center', wrap_text=wrap)
    c.border    = _borde()
    return c


def _encabezado_institucional(ws, titulo, ncols=9):
    """Cabecera verde con nombre institucional y título del reporte."""
    col_fin = get_column_letter(ncols)

    # Fila 1 — nombre institución
    ws.merge_cells(f'A1:{col_fin}1')
    c = ws['A1']
    c.value     = 'MUNICIPALIDAD PROVINCIAL DE YAU — GESTIMUNI'
    c.font      = Font(name='Calibri', bold=True, size=14, color=BLANCO)
    c.fill      = PatternFill('solid', fgColor=VERDE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Fila 2 — subtítulo
    ws.merge_cells(f'A2:{col_fin}2')
    c = ws['A2']
    c.value     = 'Sistema de Gestión Municipal con Inteligencia Artificial'
    c.font      = Font(name='Calibri', size=10, color=BLANCO, italic=True)
    c.fill      = PatternFill('solid', fgColor=VERDE_OSC)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # Fila 3 — título del reporte
    ws.merge_cells(f'A3:{col_fin}3')
    c = ws['A3']
    c.value     = titulo
    c.font      = Font(name='Calibri', bold=True, size=12, color=GRIS_HD)
    c.fill      = PatternFill('solid', fgColor=VERDE_CLR)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 22

    # Fila 4 — fecha generación
    ws.merge_cells(f'A4:{col_fin}4')
    c = ws['A4']
    c.value     = f'Generado el {datetime.now().strftime("%d de %B de %Y a las %H:%M")} hrs  |  Huánuco, Perú'
    c.font      = Font(name='Calibri', size=9, italic=True, color='FF888780')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 15

    ws.row_dimensions[5].height = 6   # espacio


def _fila_headers(ws, fila, headers, altura=20):
    for col, titulo in enumerate(headers, 1):
        c = ws.cell(row=fila, column=col, value=titulo)
        c.font      = Font(name='Calibri', bold=True, size=10, color=BLANCO)
        c.fill      = PatternFill('solid', fgColor=VERDE)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = _borde()
    ws.row_dimensions[fila].height = altura


# ══════════════════════════════════════════════════════════════════
# REPORTE DE TRÁMITES
# ══════════════════════════════════════════════════════════════════
def generar_excel_tramites(tramites, filtros=None):
    """
    Genera el Excel de trámites.
    :param tramites: lista de objetos Tramite
    :param filtros: dict con estado y prioridad (para el título)
    :return: BytesIO listo para send_file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Trámites'

    # Título dinámico
    partes = ['REPORTE DE TRÁMITES']
    if filtros:
        if filtros.get('estado'):
            partes.append(f'Estado: {filtros["estado"].upper()}')
        if filtros.get('prioridad'):
            partes.append(f'Prioridad: {filtros["prioridad"].upper()}')
    titulo = '  —  '.join(partes)

    _encabezado_institucional(ws, titulo, ncols=10)

    headers = ['N°', 'Ciudadano', 'DNI', 'Tipo de Trámite',
               'Urgencia', 'Prioridad ML', 'Confianza ML',
               'Estado', 'Observaciones', 'Fecha Registro']
    anchos  = [4, 22, 10, 24, 9, 11, 11, 12, 28, 18]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _fila_headers(ws, 6, headers)

    for idx, t in enumerate(tramites):
        fila      = 7 + idx
        alt_fill  = GRIS_ALT if idx % 2 else BLANCO
        confianza = f'{round(t.confianza_ml * 100, 1)}%'
        dni_c     = getattr(t.ciudadano, 'dni', None) or '—'

        datos = [
            idx + 1,
            t.ciudadano.nombre,
            dni_c,
            t.tipo.replace('_', ' ').title(),
            t.urgencia.upper(),
            t.prioridad_ml.upper(),
            confianza,
            t.estado.upper().replace('_', ' '),
            t.observaciones or '',
            t.fecha_registro.strftime('%d/%m/%Y %H:%M'),
        ]
        for col, valor in enumerate(datos, 1):
            fill = alt_fill
            if col == 6:   # prioridad — color semáforo
                fill = COLORES_PRIORIDAD.get(t.prioridad_ml, alt_fill)
            elif col == 8: # estado — color semáforo
                fill = COLORES_ESTADO.get(t.estado, alt_fill)
            _celda(ws, fila, col, valor,
                   negrita=(col == 1),
                   color_fill=fill,
                   alineacion='center' if col in (1, 5, 6, 7) else 'left',
                   wrap=(col == 9))
        ws.row_dimensions[fila].height = 16

    # Fila de totales
    fila_tot = 7 + len(tramites) + 1
    ws.merge_cells(f'A{fila_tot}:H{fila_tot}')
    c = ws[f'A{fila_tot}']
    c.value     = f'Total de registros: {len(tramites)}'
    c.font      = Font(name='Calibri', bold=True, size=10, color=GRIS_HD)
    c.fill      = PatternFill('solid', fgColor=VERDE_CLR)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[fila_tot].height = 18

    # Hoja resumen estadístico
    ws2 = wb.create_sheet('Resumen estadístico')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 14

    _encabezado_institucional(ws2, 'RESUMEN ESTADÍSTICO — TRÁMITES', ncols=2)
    _fila_headers(ws2, 6, ['Indicador', 'Valor'], altura=18)

    from collections import Counter
    por_estado    = Counter(t.estado    for t in tramites)
    por_prioridad = Counter(t.prioridad_ml for t in tramites)
    por_tipo      = Counter(t.tipo      for t in tramites)

    estadisticas = [
        ('Total de trámites',              len(tramites)),
        ('--- Por estado ---',             ''),
        *[(f'  {k.title().replace("_"," ")}', v) for k, v in por_estado.items()],
        ('--- Por prioridad ML ---',        ''),
        *[(f'  {k.title()}', v)            for k, v in por_prioridad.items()],
        ('--- Por tipo ---',               ''),
        *[(f'  {k.replace("_"," ").title()}', v) for k, v in por_tipo.items()],
    ]
    for i, (indicador, valor) in enumerate(estadisticas):
        fila = 7 + i
        es_header = indicador.startswith('---')
        _celda(ws2, fila, 1, indicador, negrita=es_header,
               color_fill=VERDE_CLR if es_header else (GRIS_ALT if i % 2 else BLANCO))
        _celda(ws2, fila, 2, valor,
               alineacion='center',
               color_fill=VERDE_CLR if es_header else (GRIS_ALT if i % 2 else BLANCO))
        ws2.row_dimensions[fila].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════
# REPORTE DE CURRÍCULOS / POSTULANTES
# ══════════════════════════════════════════════════════════════════
def generar_excel_curriculos(postulantes):
    """
    Genera el Excel de evaluación de currículos.
    :param postulantes: lista de objetos Postulante
    :return: BytesIO listo para send_file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Evaluación Currículos'

    _encabezado_institucional(
        ws, 'REPORTE DE EVALUACIÓN DE CURRÍCULOS — MODELO ML', ncols=12)

    headers = ['N°', 'Nombre', 'DNI', 'Puesto', 'Exp.\n(años)',
               'Educación', 'Habilidades\n(0-10)', 'Idiomas',
               'Resultado ML', 'Puntaje ML', 'Probabilidad', 'Fecha']
    anchos  = [4, 24, 10, 20, 8, 14, 12, 8, 12, 10, 12, 14]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _fila_headers(ws, 6, headers, altura=24)

    for idx, p in enumerate(postulantes):
        fila     = 7 + idx
        alt_fill = GRIS_ALT if idx % 2 else BLANCO
        color_r  = 'FFE1F5EE' if p.resultado_ml == 'apto' else 'FFFCEBEB'

        datos = [
            idx + 1,
            p.nombre,
            p.dni or '—',
            p.puesto,
            p.experiencia,
            p.educacion.title(),
            p.habilidades,
            p.idiomas or '0',
            p.resultado_ml.upper(),
            f'{p.puntaje_ml}%',
            f'{round(p.probabilidad * 100, 1)}%',
            p.fecha_registro.strftime('%d/%m/%Y'),
        ]
        for col, valor in enumerate(datos, 1):
            fill = color_r if col == 9 else alt_fill
            _celda(ws, fila, col, valor,
                   negrita=(col == 1),
                   color_fill=fill,
                   alineacion='center' if col in (1, 5, 7, 8, 9, 10, 11) else 'left')
        ws.row_dimensions[fila].height = 16

    # Totales
    aptos    = sum(1 for p in postulantes if p.resultado_ml == 'apto')
    no_aptos = len(postulantes) - aptos
    fila_res = 7 + len(postulantes) + 1
    resumen  = [
        ('Total evaluados', len(postulantes)),
        ('Aptos',           aptos),
        ('No aptos',        no_aptos),
        ('Tasa de aprobación', f'{round(aptos/len(postulantes)*100,1)}%' if postulantes else '0%'),
    ]
    for i, (lbl, val) in enumerate(resumen):
        _celda(ws, fila_res + i, 1, lbl, negrita=True,
               color_fill=VERDE_CLR)
        _celda(ws, fila_res + i, 2, val,
               alineacion='center', color_fill=VERDE_CLR)
        ws.row_dimensions[fila_res + i].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════
# REPORTE DE USUARIOS
# ══════════════════════════════════════════════════════════════════
def generar_excel_usuarios(usuarios):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Usuarios'

    _encabezado_institucional(ws, 'REPORTE DE USUARIOS DEL SISTEMA', ncols=8)

    headers = ['ID', 'Nombre', 'Usuario', 'DNI', 'Email', 'Teléfono', 'Rol', 'Estado']
    anchos  = [5, 24, 16, 10, 26, 14, 12, 10]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _fila_headers(ws, 6, headers)

    COLORES_ROL = {
        'admin':     'FFFCEBEB',
        'empleado':  'FFE6F1FB',
        'ciudadano': GRIS_ALT,
    }
    for idx, u in enumerate(usuarios):
        fila     = 7 + idx
        alt_fill = GRIS_ALT if idx % 2 else BLANCO
        datos = [
            u.id, u.nombre, u.username,
            u.dni or '—', u.email or '—', u.telefono or '—',
            u.rol.upper(),
            'ACTIVO' if u.activo else 'INACTIVO',
        ]
        for col, valor in enumerate(datos, 1):
            fill = COLORES_ROL.get(u.rol, alt_fill) if col == 7 else alt_fill
            _celda(ws, fila, col, valor,
                   color_fill=fill,
                   alineacion='center' if col in (1, 7, 8) else 'left')
        ws.row_dimensions[fila].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf